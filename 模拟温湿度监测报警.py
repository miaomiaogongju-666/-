import random
import time
import openpyxl
from openpyxl import Workbook,load_workbook

Warning_tem=60
Warning_humidity=100
wb = Workbook()
ws = wb.active
ws['A1']="温度"
ws['B1']="湿度"
i=2
while True:
    temp=random.randint(0,80)
    humidity=random.randint(0,100)
    ws.cell(row=i,column=1).value=temp
    ws.cell(row=i,column=2).value=humidity
    wb.save("recored_data.xlsx")
    i+=1
    if temp>=Warning_tem:
        print("⚠️警告！温度过高！")
    if humidity>=Warning_humidity:
        print("⚠️警告！湿度过高！")
    time.sleep(1)
    if i==50:
        break
